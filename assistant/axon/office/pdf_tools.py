"""PDF read/extract/fill: pull text, pull tables into Excel, and fill PDF form fields.
(Creating PDFs from Office files is still convert_to_pdf.)"""
import os

from openpyxl import Workbook

from axon.util import _result
from axon.office._base import _resolve_path
from axon.office.preview import open_doc, save_with_retry


def pdf_read(args):
    """Extract text from a PDF (all pages, or args.pages like '1-5'). Use this to summarize/answer
    questions about a PDF (read it, then YOU compose the summary)."""
    path = args.get("path")
    if not path or not os.path.exists(path):
        return _result(f"No such PDF: {path}", True)
    try:
        from pypdf import PdfReader
        reader = PdfReader(path)
    except Exception as e:
        return _result(f"Could not open PDF: {e}", True)
    n = len(reader.pages)
    lo, hi = 1, n
    pages = args.get("pages")
    if pages and "-" in str(pages):
        try:
            a, b = str(pages).split("-", 1)
            lo, hi = max(1, int(a)), min(n, int(b))
        except Exception:
            pass
    elif pages:
        try:
            lo = hi = max(1, min(n, int(pages)))
        except Exception:
            pass
    out = [f"{os.path.basename(path)} — {n} pages (showing {lo}-{hi})"]
    for i in range(lo - 1, hi):
        try:
            txt = (reader.pages[i].extract_text() or "").strip()
        except Exception:
            txt = ""
        out.append(f"\n--- Page {i + 1} ---\n{txt}")
    return _result("\n".join(out))


def pdf_extract_tables(args):
    """Extract tables from a PDF into an Excel workbook (one sheet per table). args: path, out?"""
    path = args.get("path")
    if not path or not os.path.exists(path):
        return _result(f"No such PDF: {path}", True)
    out = _resolve_path(args.get("out"), os.path.splitext(os.path.basename(path))[0] + " tables", ".xlsx")
    try:
        import pdfplumber
    except Exception as e:
        return _result(f"pdfplumber not available: {e}", True)
    wb = Workbook()
    wb.remove(wb.active)
    count = 0
    try:
        with pdfplumber.open(path) as pdf:
            for pi, page in enumerate(pdf.pages, 1):
                tables = page.extract_tables()  # line-ruled tables
                if not tables:                  # fall back to text alignment (borderless tables)
                    tables = page.extract_tables(
                        {"vertical_strategy": "text", "horizontal_strategy": "text"})
                for ti, table in enumerate(tables or [], 1):
                    count += 1
                    ws = wb.create_sheet(f"p{pi}_t{ti}"[:31])
                    for row in table:
                        ws.append(["" if c is None else c for c in row])
    except Exception as e:
        return _result(f"Table extraction failed: {e}", True)
    if count == 0:
        return _result("No tables found in that PDF.", True)
    wb.create_sheet("info") if not wb.sheetnames else None
    final, err = save_with_retry(out, lambda p: wb.save(p))
    if err is not None:
        return _result(f"Could not save workbook: {err}", True)
    open_doc(final)
    return _result(f"Extracted {count} table(s) into {final}.")


def pdf_fill_form(args):
    """Fill an AcroForm PDF's fields and save a copy. args: path, fields {name: value}, out?"""
    path = args.get("path")
    fields = args.get("fields") or {}
    if not path or not os.path.exists(path):
        return _result(f"No such PDF: {path}", True)
    if not fields:
        return _result("Provide fields: {field_name: value}. (Use pdf_form_fields to list names.)", True)
    out = args.get("out") or (os.path.splitext(path)[0] + " (filled).pdf")
    try:
        from pypdf import PdfReader, PdfWriter
        reader = PdfReader(path)
        writer = PdfWriter()
        writer.append(reader)
        for page in writer.pages:
            writer.update_page_form_field_values(page, {k: str(v) for k, v in fields.items()})
        with open(out, "wb") as f:
            writer.write(f)
    except Exception as e:
        return _result(f"Could not fill PDF: {e}", True)
    open_doc(out)
    return _result(f"Filled {len(fields)} field(s) -> {out}.")


def pdf_form_fields(args):
    """List the fillable field names of an AcroForm PDF (so you know what to pass to pdf_fill_form)."""
    path = args.get("path")
    if not path or not os.path.exists(path):
        return _result(f"No such PDF: {path}", True)
    try:
        from pypdf import PdfReader
        flds = PdfReader(path).get_fields() or {}
    except Exception as e:
        return _result(f"Could not read PDF fields: {e}", True)
    if not flds:
        return _result("This PDF has no fillable form fields.")
    return _result("Fields: " + ", ".join(flds.keys()))
