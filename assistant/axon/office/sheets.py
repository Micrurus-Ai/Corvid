"""Excel via openpyxl (structured editing): read a workbook's sheets/cells and apply targeted edits
(cells, formulas, ranges, sheets, formatting, charts). Saves the file, then opens it in Excel as a
live preview. Creates the workbook if it doesn't exist.
"""
import os

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

from axon.util import _result
from axon.office._base import _resolve_path
from axon.office.preview import ensure_closed, open_doc, save_with_retry

_CHARTS = {"bar": BarChart, "column": BarChart, "line": LineChart, "pie": PieChart}


def _ws(wb, op):
    name = op.get("sheet")
    if name:
        if name not in wb.sheetnames:
            return wb.create_sheet(name)
        return wb[name]
    return wb.active


def _apply_op(wb, op):
    kind = (op.get("op") or "").lower()
    if kind == "add_sheet":
        name = op.get("name") or "Sheet"
        return None if name in wb.sheetnames else (wb.create_sheet(name) and None)
    if kind == "delete_sheet":
        if op.get("name") in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb[op["name"]]
            return None
        return "sheet not found (or it's the only sheet)"
    ws = _ws(wb, op)
    if kind == "set_cell":
        ws[op["cell"]] = op.get("value")
        return None
    if kind == "set_range":
        rows = op.get("rows") or []
        col0, row0 = coordinate_from_string(op.get("start", "A1"))
        c0 = column_index_from_string(col0)
        for dr, row in enumerate(rows):
            for dc, val in enumerate(row):
                ws.cell(row=row0 + dr, column=c0 + dc, value=val)
        return None
    if kind == "format":
        rng = op.get("range")
        if not rng:
            return "format needs a range (e.g. A1:C1)"
        fnt = {}
        if op.get("bold") is not None:
            fnt["bold"] = bool(op["bold"])
        if op.get("italic") is not None:
            fnt["italic"] = bool(op["italic"])
        if op.get("font_color"):
            fnt["color"] = str(op["font_color"]).lstrip("#")
        fill = PatternFill("solid", fgColor=str(op["fill"]).lstrip("#")) if op.get("fill") else None
        align = Alignment(horizontal=op["align"]) if op.get("align") else None
        for row in ws[rng]:
            for cell in row:
                if fnt:
                    cell.font = Font(**fnt)
                if fill:
                    cell.fill = fill
                if align:
                    cell.alignment = align
                if op.get("number_format"):
                    cell.number_format = str(op["number_format"])
        return None
    if kind == "column_width":
        ws.column_dimensions[op["column"]].width = float(op.get("width", 15))
        return None
    if kind == "add_chart":
        cls = _CHARTS.get(str(op.get("chart_type", "bar")), BarChart)
        chart = cls()
        if str(op.get("chart_type")) == "column":
            chart.type = "col"
        if op.get("title"):
            chart.title = str(op["title"])
        minc, minr, maxc, maxr = range_boundaries(op["data"])
        data = Reference(ws, min_col=minc, min_row=minr, max_col=maxc, max_row=maxr)
        chart.add_data(data, titles_from_data=bool(op.get("titles_from_data", True)))
        if op.get("categories"):
            c0, r0, c1, r1 = range_boundaries(op["categories"])
            chart.set_categories(Reference(ws, min_col=c0, min_row=r0, max_col=c1, max_row=r1))
        ws.add_chart(chart, op.get("anchor", "H2"))
        return None
    return f"unknown op: {kind}"


def excel_read(args):
    """Outline a workbook: each sheet's dimensions plus a preview of the first rows/columns."""
    path = args.get("path")
    if not path or not os.path.exists(path):
        return _result(f"No such file: {path}", True)
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as e:
        return _result(f"Could not open workbook: {e}", True)
    out = [f"{os.path.basename(path)} — sheets: {', '.join(wb.sheetnames)}"]
    for name in wb.sheetnames:
        ws = wb[name]
        out.append(f"\n[{name}] {ws.max_row} rows x {ws.max_column} cols")
        for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8),
                              max_col=min(ws.max_column, 8), values_only=True):
            cells = ["" if v is None else str(v) for v in r]
            if any(cells):
                out.append("  " + " | ".join(cells))
        if ws.max_row > 8:
            out.append(f"  ... (+{ws.max_row - 8} more rows)")
    return _result("\n".join(out))


def excel_analyze(args):
    """Analyze a sheet: per numeric column compute count/sum/avg/min/max, write a 'Summary' sheet
    with a column-sums chart, and report the stats. For 'analyze this spreadsheet'. args: path, sheet?"""
    path = args.get("path")
    if not path or not os.path.exists(path):
        return _result(f"No such file: {path}", True)
    ensure_closed(path)
    try:
        wb = load_workbook(path)
    except Exception as e:
        return _result(f"Could not open workbook: {e}", True)
    name = args.get("sheet")
    ws = wb[name] if name and name in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return _result("Sheet is empty.", True)
    headers = [str(h) if h is not None else f"Col{i + 1}" for i, h in enumerate(rows[0])]
    data = rows[1:]
    stats = []
    for ci, h in enumerate(headers):
        nums = [r[ci] for r in data if ci < len(r) and isinstance(r[ci], (int, float))
                and not isinstance(r[ci], bool)]
        if nums:
            stats.append((h, len(nums), sum(nums), round(sum(nums) / len(nums), 2), min(nums), max(nums)))
    if not stats:
        return _result("No numeric columns found to analyze.", True)
    sm = wb["Summary"] if "Summary" in wb.sheetnames else wb.create_sheet("Summary")
    if sm.max_row:
        sm.delete_rows(1, sm.max_row)
    sm.append(["Column", "Count", "Sum", "Average", "Min", "Max"])
    for s in stats:
        sm.append(list(s))
    for c in sm[1]:
        c.font = Font(bold=True)
    chart = BarChart()
    chart.type = "col"
    chart.title = "Column Sums"
    chart.add_data(Reference(sm, min_col=3, min_row=1, max_row=len(stats) + 1), titles_from_data=True)
    chart.set_categories(Reference(sm, min_col=1, min_row=2, max_row=len(stats) + 1))
    sm.add_chart(chart, "H2")
    final, err = save_with_retry(path, lambda p: wb.save(p))
    if err is not None:
        return _result(f"Could not save workbook: {err}", True)
    open_doc(final)
    lines = [f"Analyzed '{ws.title}' ({len(data)} data rows). Added a Summary sheet + chart:"]
    for s in stats:
        lines.append(f"  {s[0]}: n={s[1]}, sum={s[2]}, avg={s[3]}, min={s[4]}, max={s[5]}")
    return _result("\n".join(lines))


def excel_edit(args):
    """Apply edit ops to a workbook (creating it if missing), save, and open it in Excel.

    ops example:
      [{"op":"set_range","sheet":"Q3","start":"A1","rows":[["Month","Sales"],["Jan",10],["Feb",14]]},
       {"op":"set_cell","sheet":"Q3","cell":"B4","value":"=SUM(B2:B3)"},
       {"op":"format","sheet":"Q3","range":"A1:B1","bold":true,"fill":"DDDDDD"},
       {"op":"add_chart","sheet":"Q3","chart_type":"column","data":"B1:B3","categories":"A2:A3","title":"Sales"}]
    """
    path = _resolve_path(args.get("path"), "workbook", ".xlsx")
    ops = args.get("ops") or []
    if not ops:
        return _result("Provide ops: a list of edit operations (set_cell / set_range / format / add_chart ...).", True)
    ensure_closed(path)
    try:
        wb = load_workbook(path) if os.path.exists(path) else Workbook()
    except Exception as e:
        return _result(f"Could not open workbook: {e}", True)
    done, errors = 0, []
    for n, op in enumerate(ops, 1):
        try:
            err = _apply_op(wb, op)
            if err:
                errors.append(f"op {n} ({op.get('op')}): {err}")
            else:
                done += 1
        except Exception as e:
            errors.append(f"op {n} ({op.get('op')}): {e}")
    final, serr = save_with_retry(path, lambda p: wb.save(p))
    if serr is not None:
        return _result(f"Could not save workbook: {serr}", True)
    open_doc(final)
    msg = f"Applied {done}/{len(ops)} edits to {final} (sheets: {', '.join(wb.sheetnames)})."
    if final != path:
        msg += " (Original was open/locked, so I saved an updated copy.)"
    if errors:
        msg += " Issues:\n" + "\n".join(errors)
    return _result(msg, bool(errors) and done == 0)
